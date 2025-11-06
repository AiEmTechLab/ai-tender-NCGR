# app.py — واجهة تبويبات + إصلاح KeyError + إبعاد زر التنزيل
import os, io, re, tempfile, streamlit as st, pandas as pd
from gtts import gTTS
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime

# ===== استيراد الوحدات =====
from modules.ui import setup_language, apply_theme, render_header, landing_hero
from modules.extractors import parse_criteria_from_excel, extract_text_with_pages
from modules.evaluator import evaluate_offers
from modules.analyzer import (
    suggest_criteria_from_offers,
    analyze_sections_with_pages,
    summarize_paragraphs_llm,
)
from modules.chatbot import TenderChat

# ===== إعداد الواجهة =====
T = setup_language()
apply_theme()
render_header(T)

# ===== رفع الملفات =====
if "uploaded" not in st.session_state:
    st.session_state.uploaded = False

if not st.session_state.uploaded:
    landing_hero(T)
    ex_file = st.file_uploader("📥 رفع ملف الإكسل (المعايير)", type=["xlsx", "xls"])
    offers = st.file_uploader("📥 رفع عروض الشركات (PDF/DOCX — متعدد)", type=["pdf", "docx"], accept_multiple_files=True)
    colA, colB, colC = st.columns([3, 1, 3])
    with colB:
        if st.button("🚀 ابدأ", type="primary", use_container_width=True):
            if ex_file and offers:
                st.session_state._excel = ex_file
                st.session_state._offers = offers
                st.session_state.uploaded = True
                st.rerun()
            else:
                st.warning("⚠️ فضلاً ارفع ملف المعايير والعروض أولاً.")
    st.stop()

# ===== تحميل المعايير =====
criteria_df = parse_criteria_from_excel(st.session_state._excel)
if "criteria_df" not in st.session_state:
    st.session_state.criteria_df = criteria_df.copy()
criteria_list = st.session_state.criteria_df["criterion"].tolist()

# ===== تبويبات =====
st.markdown("""
<style>
.stTabs [data-baseweb="tab-list"] {
    justify-content: center;
    gap: 25px;
    border-bottom: 2px solid #EDE9FE;
    position: sticky;
    top: 0;
    background: white;
    z-index: 1000;
    padding-top: 8px;
}
.stTabs [data-baseweb="tab"] {
    background-color: #f7f5ff;
    border-radius: 12px 12px 0 0;
    color: #5A33A4;
    font-weight: 700;
    padding: 12px 26px;
    transition: all 0.25s ease;
    font-size: 17px;
}
.stTabs [data-baseweb="tab"]:hover {
    background-color: #EDE9FE;
    transform: translateY(-2px);
}
.stTabs [aria-selected="true"] {
    background: linear-gradient(90deg, #5A33A4, #8B5CF6) !important;
    color: white !important;
    box-shadow: 0 -3px 8px rgba(90,51,164,0.25);
}
.fade-container { animation: fadeIn 0.5s ease-in-out; }
@keyframes fadeIn { 0% {opacity:0; transform:translateY(8px);} 100% {opacity:1; transform:translateY(0);} }
</style>
""", unsafe_allow_html=True)

tab1, tab2, tab3 = st.tabs([
    "⚙️ التقييم الذكي",
    "🔍📄 استكشاف المواضيع",
    "💬 المحادثة"
])

# ============================================
# ⚙️ التقييم بالذكاء الاصطناعي
# ============================================
with tab1:
    st.markdown("<div class='fade-container'>", unsafe_allow_html=True)
    st.subheader("🏆 ترتيب العروض")

    with st.expander("📋 عرض المعايير الحالية", expanded=True):
        st.dataframe(st.session_state.criteria_df, use_container_width=True)

    # 🔮 اقتراح معايير جديدة
    if st.button("🤖 اقتراح معايير جديدة من العروض"):
        st.info("🤖 جاري تحليل العروض واقتراح معايير جديدة...")
        offers_texts = []
        for f in st.session_state._offers:
            data = extract_text_with_pages(f)
            if isinstance(data, dict):
                if data.get("type") == "pdf":
                    offers_texts.append("\n".join([p["text"] for p in data.get("pages", [])]))
                elif data.get("type") == "docx":
                    offers_texts.append(data.get("text", ""))
        suggested_all = suggest_criteria_from_offers(offers_texts, criteria_list) or []
        suggested_all = suggested_all[:5]
        synonyms = {s: [s, s.replace(" ", "_"), s.lower()] for s in suggested_all}
        results = []
        for s, syns in synonyms.items():
            count, pages_found = 0, []
            for f in st.session_state._offers:
                data = extract_text_with_pages(f)
                if isinstance(data, dict) and "pages" in data:
                    for p in data["pages"]:
                        text = p["text"]
                        for term in syns:
                            if re.search(rf"\b{re.escape(term)}\b", text, re.IGNORECASE):
                                count += 1
                                pages_found.append(p["page_num"])
            pages_str = ", ".join(map(str, sorted(set(pages_found)))) or "-"
            weight = min(5, 1 + count // 3)
            results.append({
                "criterion": s,
                "synonyms": ", ".join(syns),
                "count": count,
                "pages": pages_str,
                "weight": weight
            })
        st.session_state.suggested_criteria_df = pd.DataFrame(results)
        st.success("✅ تم توليد معايير جديدة بنجاح!")

    # عرض المقترحات وإضافتها
    if "suggested_criteria_df" in st.session_state and not st.session_state.suggested_criteria_df.empty:
        df = st.session_state.suggested_criteria_df
        st.dataframe(df, use_container_width=True)
        selected = st.multiselect("حدد المعايير التي تريد إضافتها:", options=df["criterion"].tolist())
        if selected and st.button("📥 إضافة المحدد وبدء التقييم"):
            to_add = df[df["criterion"].isin(selected)][["criterion", "weight"]]
            st.session_state.criteria_df = pd.concat(
                [st.session_state.criteria_df, to_add], ignore_index=True
            ).drop_duplicates(subset=["criterion"], keep="last")
            ranked, details = evaluate_offers(st.session_state._offers, st.session_state.criteria_df["criterion"].tolist())
            st.session_state.results = ranked
            st.session_state.details = details
            st.success("✅ تم تشغيل التقييم!")
            st.rerun()

    # تشغيل التقييم مباشرة
    if st.button("⚙️ تشغيل التقييم الذكي", type="primary"):
        ranked, details = evaluate_offers(st.session_state._offers, criteria_list)
        st.session_state.results = ranked
        st.session_state.details = details
        st.success("✅ تم اكتمال التقييم!")
        st.rerun()

    # عرض النتائج والتفسير
    if "results" in st.session_state:
        ranked = st.session_state.results.copy()
        details = st.session_state.details
        ranked["النسبة %"] = (ranked["overall"] * 100).round(1)
        st.dataframe(ranked[["file", "النسبة %"]], use_container_width=True)
        best = ranked.iloc[0]
        st.markdown(f"✅ **أفضل عرض:** {best['file']} بنسبة {best['النسبة %']}%")

        # 🔍 تفسير الذكاء الصناعي
        explanation = ""
        try:
            from groq import Groq
            client = Groq(api_key=os.getenv("GROQ_API_KEY"))
            prompt = f"بناءً على النتائج التالية:\n{ranked.to_string(index=False)}\nاشرح بالعربية المختصرة لماذا العرض {best['file']} هو الأفضل."
            resp = client.chat.completions.create(
                model="llama-3.3-70b-versatile",
                messages=[{"role": "user", "content": prompt}],
                temperature=0.4,
            )
            explanation = resp.choices[0].message.content.strip()
            st.markdown("### 🧾 سبب اختيار العرض الأفضل")
            st.markdown(
                f"<div style='background:#f5f0ff;border-right:5px solid #5A33A4;padding:15px;border-radius:10px;text-align:justify;margin-bottom:25px;'>{explanation}</div>",
                unsafe_allow_html=True
            )
        except Exception as e:
            st.warning(f"⚠️ لم يتمكن النظام من توليد التفسير: {e}")

        # 📊 زر تنزيل التقرير الكامل (Excel)
        if st.button("📊 تنزيل التقرير الكامل (Excel)"):
            PURPLE_DARK = "4B2E83"
            PURPLE_LIGHT = "8B5CF6"
            ROW_ALT = "F5F0FF"
            WHITE = "FFFFFF"

            wb = Workbook()
            ws_rank = wb.active
            ws_rank.title = "🏆 الترتيب النهائي"

            # عنوان
            ws_rank.merge_cells("A1:C1")
            ws_rank["A1"] = "📊 الترتيب النهائي للعروض"
            ws_rank["A1"].font = Font(bold=True, size=16, color=WHITE)
            ws_rank["A1"].alignment = Alignment(horizontal="center", vertical="center")
            ws_rank["A1"].fill = PatternFill(start_color=PURPLE_DARK, end_color=PURPLE_DARK, fill_type="solid")

            ws_rank.append(["اسم العرض", "النسبة %", "الدرجة الإجمالية"])
            for cell in ws_rank[2]:
                cell.font = Font(bold=True, color=WHITE)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.fill = PatternFill(start_color=PURPLE_LIGHT, end_color=PURPLE_LIGHT, fill_type="solid")

            for i, (_, row) in enumerate(ranked.iterrows(), start=0):
                percent = round(row["overall"] * 100, 1)
                ws_rank.append([row["file"], percent, round(row["overall"], 3)])
                if i % 2 == 0:
                    for c in ws_rank[ws_rank.max_row]:
                        c.fill = PatternFill(start_color=ROW_ALT, end_color=ROW_ALT, fill_type="solid")
                for c in ws_rank[ws_rank.max_row]:
                    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    c.font = Font(size=13)

            for col_idx in range(1, ws_rank.max_column + 1):
                ws_rank.column_dimensions[get_column_letter(col_idx)].width = 35

            # 🧾 تفاصيل العروض
            for fname, df in details.items():
                ws = wb.create_sheet(title=fname[:28])
                ws.merge_cells("A1:E1")
                ws["A1"] = f"📋 تفاصيل العرض: {fname}"
                ws["A1"].font = Font(bold=True, size=15, color=WHITE)
                ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
                ws["A1"].fill = PatternFill(start_color=PURPLE_DARK, end_color=PURPLE_DARK, fill_type="solid")

                ws.append(["المعيار", "الدرجة", "تحويل (0..1)", "السبب", "سؤال الذكاء الصناعي"])
                for cell in ws[2]:
                    cell.font = Font(bold=True, color=WHITE)
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    cell.fill = PatternFill(start_color=PURPLE_LIGHT, end_color=PURPLE_LIGHT, fill_type="solid")

                df2 = df.copy()
                # ✅ إصلاح الأعمدة المفقودة
                if "reason" not in df2.columns:
                    df2["reason"] = ""
                if "ai_question" not in df2.columns:
                    df2["ai_question"] = ""
                if "score" not in df2.columns:
                    df2["score"] = 0

                df2["reason"] = df2["reason"].astype(str)
                df2["ai_question"] = df2["ai_question"].astype(str)
                df2["score"] = df2["score"].astype(float)
                df2["تحويل (0..1)"] = ((df2["score"] - 1) / 3).round(3)

                for i, r in enumerate(df2.itertuples(), start=0):
                    ws.append([
                        r.criterion,
                        r.score,
                        r._asdict().get("تحويل (0..1)", ""),
                        r.reason,
                        r.ai_question
                    ])
                    if i % 2 == 0:
                        for c in ws[ws.max_row]:
                            c.fill = PatternFill(start_color=ROW_ALT, end_color=ROW_ALT, fill_type="solid")
                    for c in ws[ws.max_row]:
                        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                        c.font = Font(size=13)

                for col_idx in range(1, ws.max_column + 1):
                    ws.column_dimensions[get_column_letter(col_idx)].width = 45

            if explanation:
                ws_exp = wb.create_sheet("🧠 سبب الاختيار")
                ws_exp["A1"] = "🧠 سبب اختيار العرض الأفضل"
                ws_exp["A1"].font = Font(bold=True, size=15, color=WHITE)
                ws_exp["A1"].alignment = Alignment(horizontal="center", vertical="center")
                ws_exp["A1"].fill = PatternFill(start_color=PURPLE_DARK, end_color=PURPLE_DARK, fill_type="solid")
                ws_exp["A2"] = explanation
                ws_exp["A2"].alignment = Alignment(wrap_text=True, vertical="top")
                ws_exp.column_dimensions["A"].width = 100

            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            st.download_button(
                "⬇️ تحميل التقرير الكامل (Excel)",
                data=buffer,
                file_name=f"SmartTender_Report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

    st.markdown("</div>", unsafe_allow_html=True)



# ============================================
# 🔍📄 استكشاف المواضيع (تحليل الأقسام والملخصات)
# ============================================
with tab2:
    st.markdown("<div class='fade-container'>", unsafe_allow_html=True)
    st.subheader("🧭 تحليل العروض الفنية واستخراج الأقسام والملخصات")

    # خطوة 1: تحليل العروض تلقائيًا
    if st.button("🔍 تحليل العروض تلقائيًا"):
        st.info("🤖 جاري قراءة العروض واستخراج الأقسام...")
        topics_data = {}
        for offer in st.session_state._offers:
            try:
                st.markdown(f"📂 **جارٍ تحليل العرض:** {offer.name}")
                data = extract_text_with_pages(offer)
                sections = analyze_sections_with_pages(data)
                topics_data[offer.name] = sections
                st.success(f"✅ تم تحليل {offer.name} بنجاح ({len(sections)} قسم).")
            except Exception as e:
                st.error(f"⚠️ خطأ أثناء تحليل {offer.name}: {e}")
        st.session_state.topics = topics_data

    # خطوة 2: عرض النتائج
    if "topics" in st.session_state and st.session_state.topics:
        offers_names = list(st.session_state.topics.keys())
        selected_offer = st.selectbox("📘 اختر عرضًا:", offers_names)
        if selected_offer:
            sections = st.session_state.topics[selected_offer]
            # اختيار قسم
            names = [f"{s['section']} (📄 صفحة {s['start_page']})" for s in sections]
            label = st.selectbox("📄 اختر قسمًا:", names)
            if label:
                selected_section = label.split(" (")[0]
                sec = next((s for s in sections if s["section"] == selected_section), None)
                if sec:
                    st.markdown(f"### 🟣 {sec['section']}")
                    st.markdown(f"**📄 يبدأ من الصفحة:** {sec['start_page']}")
                    st.markdown(f"**📝 الملخص المبدئي:** {sec['summary']}")
                    st.divider()
                    # نص القسم الكامل ضمن Expander
                    with st.expander("📜 عرض النص الكامل للقسم", expanded=False):
                        st.markdown(
                            f"<div style='background:#f9f9f9;padding:14px;border-radius:10px;text-align:justify;white-space:pre-wrap;'>{sec['content']}</div>",
                            unsafe_allow_html=True
                        )
                    # تلخيص تفصيلي LLM
                    if st.button("🪄 توليد ملخص تفصيلي للقسم", key=f"summ_{selected_offer}_{selected_section}"):
                        try:
                            summary = summarize_paragraphs_llm(sec["content"][:18000])
                            st.success("✅ تم توليد الملخص بنجاح!")
                            st.markdown("### ✨ الملخص الذكي")
                            st.markdown(
                                f"<div style='background:#f5f0ff;border-right:5px solid #5A33A4;padding:15px;border-radius:10px;text-align:justify;'>{summary}</div>",
                                unsafe_allow_html=True
                            )
                            st.download_button(
                                "⬇️ تنزيل النص والملخص",
                                data=f"القسم: {sec['section']}\nالصفحة: {sec['start_page']}\n\nالنص:\n{sec['content']}\n\nالملخص:\n{summary}".encode("utf-8"),
                                file_name=f"{selected_offer}_{sec['section']}_summary.txt",
                                mime="text/plain",
                                use_container_width=True
                            )
                        except Exception as e:
                            st.error(f"⚠️ لم يتم توليد الملخص: {e}")
    st.markdown("</div>", unsafe_allow_html=True)

# ============================================
# 💬 المحادثة (شاتبوت)
# ============================================
with tab3:
    st.markdown("<div class='fade-container'>", unsafe_allow_html=True)
    st.subheader("💬 الشاتبوت الذكي للعروض المرفوعة")
    st.markdown("<p style='color:#666;'>اختر العرض الذي ترغب في مناقشته، وسيجيب المساعد بالعربية مع ذكر رقم الصفحة عند الإمكان.</p>", unsafe_allow_html=True)

    offers_names = [f.name for f in st.session_state._offers]
    selected_offer = st.selectbox("📂 اختر عرضًا:", offers_names)
    if not selected_offer:
        st.info("📋 يرجى اختيار عرض أولًا لبدء المحادثة.")
        st.stop()

    # بناء سياق النص (مع أرقام الصفحات)
    if "chat_ctx" not in st.session_state or st.session_state.get("ctx_name") != selected_offer:
        for f in st.session_state._offers:
            if f.name == selected_offer:
                try:
                    data = extract_text_with_pages(f)
                    if isinstance(data, dict):
                        ctx_text = "\n".join([f"[صفحة {p['page_num']}]\n{p['text']}" for p in data.get("pages", [])])
                        st.session_state.chat_ctx = {selected_offer: ctx_text}
                        st.session_state.ctx_name = selected_offer
                        f.seek(0)
                except Exception as e:
                    st.error(f"⚠️ لم يتمكن من قراءة {selected_offer}: {e}")
                    st.stop()
        st.session_state.chat_msgs = []
        st.session_state.chatbot = TenderChat(st.session_state.chat_ctx)

    # عرض سجل المحادثة
    chat_html = "<div style='display:flex;flex-direction:column;gap:6px;margin-bottom:10px;'>"
    for role, msg in st.session_state.chat_msgs:
        if role == "user":
            chat_html += f"<div style='background:#E5E7EB;color:#111827;padding:10px 14px;border-radius:16px;align-self:flex-end;max-width:80%;'>{msg}</div>"
        else:
            chat_html += f"<div style='background:#5A33A4;color:white;padding:10px 14px;border-radius:16px;align-self:flex-start;max-width:80%;'>{msg}</div>"
    chat_html += "</div>"
    st.markdown(chat_html, unsafe_allow_html=True)

    # إدخال المستخدم والرد
    st.markdown("<div style='height:100px'></div>", unsafe_allow_html=True)
    user_input = st.chat_input(f"💭 اكتب سؤالك عن {selected_offer}...")
    if user_input:
        st.session_state.chat_msgs.append(("user", user_input))
        with st.spinner("🤖 المساعد يكتب الآن..."):
            answer = st.session_state.chatbot.answer(
                f"العرض الحالي هو: {selected_offer}\n\nالسؤال: {user_input}"
            )
        # محاولة التقاط رقم الصفحة
        _re = re
        m = _re.search(r"صفحة\s+(\d+)", answer)
        if m:
            answer += f"<br><br>📄 <i>المعلومة وردت في الصفحة رقم {m.group(1)}.</i>"
        answer += f"<br><br>🗂️ <i>الإجابة مستندة إلى عرض:</i> <b>{selected_offer}</b>"
        st.session_state.chat_msgs.append(("assistant", answer))
        # صوت عربي
        try:
            tts = gTTS(text=answer, lang='ar')
            with tempfile.NamedTemporaryFile(delete=False, suffix=".mp3") as tmp:
                tts.save(tmp.name)
                st.audio(tmp.name, format="audio/mp3", start_time=0)
        except Exception as e:
            st.warning(f"⚠️ لم يتمكن من توليد الصوت: {e}")
        st.rerun()
    st.markdown("</div>", unsafe_allow_html=True)
